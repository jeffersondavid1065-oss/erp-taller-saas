import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from sqlalchemy import text
from db import obtener_conexion, mensaje_error_amigable
from queries import (
    obtener_categorias_gasto,
    obtener_gastos_filtrado,
    obtener_gastos_por_categoria,
    obtener_metricas_financieras,
    invalidar_cache_gastos,
)
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Gastos y Análisis Financiero", layout="wide")

st.markdown("""
    <style>
    header::after {
        content: "";
        position: fixed !important;
        top: 0 !important;
        right: 0 !important;
        width: 350px !important;
        height: 60px !important;
        background-color: var(--background-color) !important;
        z-index: 9999999 !important;
        pointer-events: all !important;
    }
    @keyframes fade-in-up {
        0% { opacity: 0; transform: translateY(20px); }
        100% { opacity: 1; transform: translateY(0); }
    }
    [data-testid="stAppViewBlockContainer"] { animation: fade-in-up 0.15s ease-out; }
    </style>
""", unsafe_allow_html=True)

# Autenticación
if "auth" not in st.session_state:
    st.session_state.auth = {"logged": False, "user_id": None, "nombre_taller": None, "email": None}

if not st.session_state.auth["logged"]:
    st.warning("Debes iniciar sesión para acceder a este módulo.")
    st.stop()
    
    # Bloqueo de rol: los operarios de Patio solo tienen acceso a Recepción.
if st.session_state.auth.get("rol") == "patio":
    st.warning("🔒 Tu usuario solo tiene acceso al módulo de Recepción de Vehículos.")
    st.stop()

engine = obtener_conexion()
user_id = st.session_state.auth["user_id"]
nombre_taller = st.session_state.auth["nombre_taller"]

def formato_cop(numero):
    return f"${float(numero):,.0f}".replace(",", ".")


def generar_excel_gastos(df_gastos, titulo_reporte, nombre_archivo):
    """Genera un archivo Excel formateado profesionalmente para reportes de gastos."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos"

    # Estilos
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Título
    ws.merge_cells("A1:G1")
    ws["A1"] = titulo_reporte
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    # Subtítulo con fecha
    ws.merge_cells("A2:G2")
    ws["A2"] = f"Generado: {datetime.today().strftime('%d de %B de %Y')}"
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.append([])  # Fila en blanco

    # Headers
    headers = ["Fecha", "Categoría", "Descripción", "Tipo", "Monto ($)"]
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Datos
    for idx, row in df_gastos.iterrows():
        ws.append([
            row['fecha'],
            row['categoria'],
            row['descripcion'],
            row['tipo'],
            row['monto']
        ])
        
        # Formato de datos
        for col_num in range(1, 6):
            cell = ws.cell(row=5 + idx, column=col_num)
            cell.border = border
            if col_num == 5:  # Columna de montos
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    # Fila de totales
    total_row = len(df_gastos) + 5
    ws.merge_cells(f"A{total_row}:D{total_row}")
    ws[f"A{total_row}"] = "TOTAL GASTOS"
    ws[f"A{total_row}"].font = total_font
    ws[f"A{total_row}"].fill = total_fill
    ws[f"A{total_row}"].alignment = Alignment(horizontal="right")
    ws[f"A{total_row}"].border = border

    total_monto = df_gastos['monto'].sum()
    total_cell = ws[f"E{total_row}"]
    total_cell.value = total_monto
    total_cell.font = total_font
    total_cell.fill = total_fill
    total_cell.number_format = '#,##0.00'
    total_cell.alignment = Alignment(horizontal="right")
    total_cell.border = border

    # Ajustar ancho de columnas
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 15

    # Guardar a BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer


def generar_excel_analisis_mensual(ingresos, gastos, margen, df_gastos_mes, año, mes, taller_nombre):
    """Genera un reporte financiero mensual completo en Excel."""
    wb = Workbook()
    
    # Hoja 1: Resumen
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Título
    ws_resumen.merge_cells("A1:D1")
    ws_resumen["A1"] = f"REPORTE FINANCIERO MENSUAL - {taller_nombre}"
    ws_resumen["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws_resumen["A1"].alignment = Alignment(horizontal="center")
    
    ws_resumen.merge_cells("A2:D2")
    meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    ws_resumen["A2"] = f"{meses[mes-1]} de {año}"
    ws_resumen["A2"].font = Font(italic=True, size=11, color="666666")
    ws_resumen["A2"].alignment = Alignment(horizontal="center")
    
    ws_resumen.append([])
    ws_resumen.append([])
    
    # Métricas
    ws_resumen.append(["CONCEPTO", "VALOR"])
    ws_resumen["A5"].font = title_font
    ws_resumen["A5"].fill = title_fill
    ws_resumen["B5"].font = title_font
    ws_resumen["B5"].fill = title_fill
    
    ws_resumen.append(["Ingresos Facturados", ingresos])
    ws_resumen.append(["Gastos Totales", gastos])
    ws_resumen.append(["Margen Neto", margen])
    
    # Formato
    for row in range(6, 9):
        ws_resumen.cell(row=row, column=1).font = Font(bold=True)
        ws_resumen.cell(row=row, column=2).number_format = '#,##0.00'
        ws_resumen.cell(row=row, column=2).alignment = Alignment(horizontal="right")
    
    if margen < 0:
        ws_resumen.cell(row=8, column=2).font = Font(bold=True, color="FF0000")
    else:
        ws_resumen.cell(row=8, column=2).font = Font(bold=True, color="00B050")
    
    ws_resumen.column_dimensions["A"].width = 25
    ws_resumen.column_dimensions["B"].width = 18
    
    # Hoja 2: Detalle de Gastos
    ws_detalle = wb.create_sheet("Detalle Gastos")
    
    ws_detalle.append(["DETALLE DE GASTOS"])
    ws_detalle.merge_cells("A1:E1")
    ws_detalle["A1"].font = Font(bold=True, size=12, color="1F4E78")
    ws_detalle.append([])
    
    headers = ["Fecha", "Categoría", "Descripción", "Tipo", "Monto ($)"]
    ws_detalle.append(headers)
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_detalle.cell(row=3, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    for idx, row in df_gastos_mes.iterrows():
        ws_detalle.append([
            row['fecha'],
            row['categoria'],
            row['descripcion'],
            row['tipo'],
            row['monto']
        ])
        
        for col_num in range(1, 6):
            cell = ws_detalle.cell(row=4 + idx, column=col_num)
            if col_num == 5:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
    
    ws_detalle.column_dimensions["A"].width = 15
    ws_detalle.column_dimensions["B"].width = 20
    ws_detalle.column_dimensions["C"].width = 35
    ws_detalle.column_dimensions["D"].width = 12
    ws_detalle.column_dimensions["E"].width = 15
    
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

st.title("Gastos y Análisis Financiero")
st.markdown(f"Control de gastos operacionales para: **{nombre_taller}**")
st.markdown("---")

# Crear categorías predefinidas si no existen
@st.cache_data(ttl=3600)
def inicializar_categorias_predefinidas(uid):
    with engine.begin() as conn:
        # Verificar si ya existen categorías para este usuario
        existe = conn.execute(text("SELECT COUNT(*) FROM Categorias_Gasto WHERE usuario_id = :uid"), {"uid": uid}).scalar()
        if existe == 0:
            categorias_default = [
                ("Nómina Mecánicos", "Gastos de personal"),
                ("Arriendo/Local", "Renta del local del taller"),
                ("Servicios", "Agua, luz, teléfono, internet"),
                ("Repuestos/Insumos", "Compra de materiales"),
                ("Combustible/Transporte", "Gasolina y transporte"),
                ("Herramientas", "Compra o mantenimiento de herramientas"),
                ("Mantenimiento Equipos", "Reparación de maquinaria"),
                ("Impuestos/Licencias", "Pagos tributarios"),
                ("Seguros", "Pólizas de seguro"),
                ("Otros", "Gastos varios"),
            ]
            for nombre, desc in categorias_default:
                conn.execute(
                    text("INSERT INTO Categorias_Gasto (usuario_id, nombre, descripcion, tipo) VALUES (:uid, :nom, :desc, 'Variable')"),
                    {"uid": uid, "nom": nombre, "desc": desc}
                )
    return True

inicializar_categorias_predefinidas(user_id)

tab_registrar, tab_historial, tab_analisis = st.tabs(["Registrar Gasto", "Historial de Gastos", "Margen y Rentabilidad"])

# ==========================================
# TAB 1: REGISTRAR GASTO
# ==========================================
with tab_registrar:
    categorias = obtener_categorias_gasto(user_id)
    if categorias:
        dict_categorias = {c[1]: c[0] for c in categorias}
        opciones_categorias = list(dict_categorias.keys())

        with st.form("form_nuevo_gasto", clear_on_submit=True):
            col1, col2 = st.columns(2)
            with col1:
                categoria_sel = st.selectbox("Categoría de Gasto", options=opciones_categorias)
                descripcion = st.text_input("Descripción del gasto")
                monto = st.number_input("Monto ($)", min_value=0.0, step=1000.0)
            with col2:
                fecha_gasto = st.date_input("Fecha del gasto", value=datetime.today())
                tipo_gasto = st.selectbox("Tipo", ["Variable", "Fijo"])

            st.markdown("")
            if st.form_submit_button("Guardar Gasto", type="primary"):
                if descripcion and monto > 0 and categoria_sel:
                    try:
                        with engine.begin() as conn:
                            conn.execute(
                                text("""
                                    INSERT INTO Gastos (usuario_id, categoria_id, descripcion, monto, fecha, tipo)
                                    VALUES (:uid, :cat_id, :desc, :monto, :fecha, :tipo)
                                """),
                                {
                                    "uid": user_id,
                                    "cat_id": dict_categorias[categoria_sel],
                                    "desc": descripcion,
                                    "monto": float(monto),
                                    "fecha": fecha_gasto.strftime('%Y-%m-%d'),
                                    "tipo": tipo_gasto
                                }
                            )
                        invalidar_cache_gastos()
                        st.success(f"Gasto de {formato_cop(monto)} registrado en {categoria_sel}.")
                        st.rerun()
                    except Exception as e:
                        st.error(mensaje_error_amigable(e, "guardar el gasto"))
                else:
                    st.warning("Completa todos los campos obligatorios.")
    else:
        st.info("Inicializando categorías de gasto...")
        st.rerun()

# ==========================================
# TAB 2: HISTORIAL DE GASTOS
# ==========================================
with tab_historial:
    st.subheader("Historial de Gastos Registrados")

    col_f1, col_f2, col_f3 = st.columns(3)
    with col_f1:
        hoy = datetime.today()
        hace_30 = hoy - timedelta(days=30)
        fechas = st.date_input("Rango de fechas", [hace_30, hoy])
    with col_f2:
        categorias = obtener_categorias_gasto(user_id)
        dict_categorias_filtro = {c[1]: c[0] for c in categorias}
        filtro_categoria = st.selectbox("Filtrar por categoría (opcional)", ["-- Todas --"] + list(dict_categorias_filtro.keys()))
    with col_f3:
        st.write("")  # Espaciador

    if len(fechas) == 2:
        fecha_ini, fecha_fin = fechas
        categoria_id_filtro = None if filtro_categoria == "-- Todas --" else dict_categorias_filtro[filtro_categoria]

        df_gastos = obtener_gastos_filtrado(user_id, fecha_ini, fecha_fin, categoria_id_filtro)

        if not df_gastos.empty:
            total_periodo = df_gastos['monto'].sum()
            st.success(f"Total gastos en el período: {formato_cop(total_periodo)}")

            df_mostrar = df_gastos[['categoria', 'descripcion', 'monto', 'fecha', 'tipo']].copy()
            df_mostrar.columns = ['Categoría', 'Descripción', 'Monto ($)', 'Fecha', 'Tipo']

            st.dataframe(
                df_mostrar.style.format({'Monto ($)': lambda x: formato_cop(x)}),
                use_container_width=True,
                hide_index=True
            )

            st.markdown("---")
            
            col_descargar, col_resumen = st.columns([2, 2])
            with col_descargar:
                # Descargar en Excel
                excel_buffer = generar_excel_gastos(
                    df_gastos,
                    f"Listado de Gastos - {fecha_ini.strftime('%d/%m/%Y')} a {fecha_fin.strftime('%d/%m/%Y')}",
                    f"Gastos_{fecha_ini}_{fecha_fin}.xlsx"
                )
                st.download_button(
                    label="📥 Descargar en Excel (para DIAN)",
                    data=excel_buffer,
                    file_name=f"Gastos_{fecha_ini}_{fecha_fin}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            with col_resumen:
                st.markdown("**Resumen Rápido:**")
                df_fijos_vars = df_gastos.groupby('tipo')['monto'].sum().reset_index()
                for idx, row in df_fijos_vars.iterrows():
                    st.write(f"• {row['tipo']}: {formato_cop(row['monto'])}")

            st.markdown("---")
            
            col_exp1, col_exp2 = st.columns(2)
            with col_exp1:
                # Top 3 categorías
                df_top_cat = df_gastos.groupby('categoria')['monto'].sum().nlargest(3).reset_index()
                st.markdown("**Top 3 Categorías por Gasto:**")
                for idx, row in df_top_cat.iterrows():
                    st.write(f"• {row['categoria']}: {formato_cop(row['monto'])}")
            with col_exp2:
                st.write("")
        else:
            st.info("No hay gastos registrados en este período.")

# ==========================================
# TAB 3: ANÁLISIS FINANCIERO
# ==========================================
with tab_analisis:
    st.subheader("Margen y Rentabilidad del Mes")

    col_mes1, col_mes2 = st.columns(2)
    with col_mes1:
        mes_sel = st.selectbox("Mes", range(1, 13), format_func=lambda m: ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"][m-1])
    with col_mes2:
        año_sel = st.number_input("Año", value=datetime.today().year, min_value=2020)

    fecha_inicio_mes = datetime(año_sel, mes_sel, 1).date()
    if mes_sel == 12:
        fecha_fin_mes = datetime(año_sel + 1, 1, 1).date() - timedelta(days=1)
    else:
        fecha_fin_mes = datetime(año_sel, mes_sel + 1, 1).date() - timedelta(days=1)

    df_gastos_mes = obtener_gastos_filtrado(user_id, fecha_inicio_mes, fecha_fin_mes)

    # Ingresos y gastos del mes ya sin el IVA cobrado mezclado (el IVA no es
    # ingreso del taller - ver la página "Análisis Financiero" para el detalle a declarar).
    ingresos_totales, gastos_totales = obtener_metricas_financieras(user_id, año_sel, mes_sel)

    margen_neto = ingresos_totales - gastos_totales

    col_m1, col_m2, col_m3, col_m4 = st.columns(4)
    col_m1.metric("Ingresos", formato_cop(ingresos_totales))
    col_m2.metric("Gastos", formato_cop(gastos_totales), delta=f"-{formato_cop(gastos_totales)}")
    col_m3.metric("Margen Neto", formato_cop(margen_neto), delta_color="inverse" if margen_neto < 0 else "normal")
    if ingresos_totales > 0:
        pct_margen = (margen_neto / ingresos_totales) * 100
        col_m4.metric("% Margen", f"{pct_margen:.1f}%")
    st.caption("Los ingresos no incluyen el IVA cobrado (ese IVA hay que declararlo, no es utilidad del taller). Ve al módulo **Análisis Financiero** en el menú lateral para ver el detalle a declarar.")

    st.markdown("---")

    col_graf1, col_graf2 = st.columns(2)

    with col_graf1:
        st.markdown("**Gastos por Categoría**")
        df_por_cat = obtener_gastos_por_categoria(user_id, fecha_inicio_mes, fecha_fin_mes)
        if not df_por_cat.empty:
            st.bar_chart(df_por_cat.set_index('nombre')['total'], height=300)
        else:
            st.info("Sin gastos en este mes.")

    with col_graf2:
        st.markdown("**Proporción Gastos Fijos vs Variables**")
        if not df_gastos_mes.empty:
            df_fijos_vars_pie = df_gastos_mes.groupby('tipo')['monto'].sum().reset_index()
            if not df_fijos_vars_pie.empty:
                st.bar_chart(df_fijos_vars_pie.set_index('tipo')['monto'], height=300)
            else:
                st.info("Sin gastos en este mes.")
        else:
            st.info("Sin gastos en este mes.")

    st.markdown("---")
    st.markdown("**Resumen Detallado**")
    
    col_res1, col_res2, col_res3 = st.columns(3)
    with col_res1:
        st.write("**Ingresos Facturados:**")
        st.write(f"`{formato_cop(ingresos_totales)}`")
    with col_res2:
        st.write("**Total Gastos:**")
        st.write(f"`{formato_cop(gastos_totales)}`")
    with col_res3:
        st.write("**Rentabilidad (Margen Neto):**")
        color = "🔴" if margen_neto < 0 else "🟢"
        st.write(f"{color} `{formato_cop(margen_neto)}`")

    st.markdown("---")
    
    col_desc1, col_desc2 = st.columns(2)
    with col_desc1:
        # Descargar reporte mensual completo
        excel_reporte = generar_excel_analisis_mensual(
            ingresos_totales,
            gastos_totales,
            margen_neto,
            df_gastos_mes,
            año_sel,
            mes_sel,
            nombre_taller
        )
        st.download_button(
            label="📊 Descargar Reporte Mensual Completo",
            data=excel_reporte,
            file_name=f"Reporte_Financiero_{mes_sel:02d}_{año_sel}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    
    with col_desc2:
        st.info("💡 **Tip para la DIAN:** Descarga estos reportes mensuales para mantener un registro organizado de tus ingresos y gastos. Facilita auditorías y declaraciones de impuestos.")
