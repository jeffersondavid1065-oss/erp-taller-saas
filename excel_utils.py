"""
Reportes Excel con formato profesional, para que el dueño del taller se los
pueda mandar directo a su contador. Sigue el mismo lenguaje visual ya usado
en Gastos y análisis (encabezado azul oscuro, bordes finos, fila de total
resaltada), centralizado acá para reutilizarlo en otras páginas.
"""
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
_TOTAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_TOTAL_FONT = Font(bold=True, size=11)
_TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
_SUBTITLE_FONT = Font(italic=True, size=10, color="666666")
_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def generar_excel_tabla(df, titulo_reporte, nombre_taller, columnas_moneda=None,
                         nombre_hoja="Datos", columna_total=None):
    """
    Convierte un DataFrame en un Excel con formato profesional: título
    centrado, encabezados con fondo azul oscuro, bordes finos, columnas de
    moneda formateadas como $, ancho de columna ajustado al contenido y fila
    de total opcional. Devuelve los bytes del archivo (listos para
    st.download_button).

    columnas_moneda: nombres de columna (tal como aparecen en df.columns)
    que se deben formatear como moneda.
    columna_total: si se indica, agrega una fila final "TOTAL" sumando esa columna.
    """
    columnas_moneda = columnas_moneda or []
    n_cols = len(df.columns)

    wb = Workbook()
    ws = wb.active
    ws.title = nombre_hoja[:31]

    ultima_col_letra = get_column_letter(n_cols)

    ws.merge_cells(f"A1:{ultima_col_letra}1")
    ws["A1"] = titulo_reporte
    ws["A1"].font = _TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    ws.merge_cells(f"A2:{ultima_col_letra}2")
    ws["A2"] = f"{nombre_taller} — Generado el {datetime.today().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = _SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="center")

    fila_header = 4
    for col_idx, nombre_col in enumerate(df.columns, start=1):
        celda = ws.cell(row=fila_header, column=col_idx, value=str(nombre_col))
        celda.font = _HEADER_FONT
        celda.fill = _HEADER_FILL
        celda.alignment = Alignment(horizontal="center", vertical="center")
        celda.border = _BORDER

    fila_actual = fila_header + 1
    for _, fila in df.iterrows():
        for col_idx, nombre_col in enumerate(df.columns, start=1):
            valor = fila[nombre_col]
            celda = ws.cell(row=fila_actual, column=col_idx, value=valor)
            celda.border = _BORDER
            if nombre_col in columnas_moneda:
                celda.number_format = '"$"#,##0'
                celda.alignment = Alignment(horizontal="right")
            else:
                celda.alignment = Alignment(horizontal="left")
        fila_actual += 1

    if columna_total and columna_total in df.columns and not df.empty:
        idx_col_total = list(df.columns).index(columna_total) + 1
        if idx_col_total > 1:
            ws.merge_cells(
                start_row=fila_actual, start_column=1,
                end_row=fila_actual, end_column=idx_col_total - 1
            )
        celda_etiqueta = ws.cell(row=fila_actual, column=1, value="TOTAL")
        celda_etiqueta.font = _TOTAL_FONT
        celda_etiqueta.fill = _TOTAL_FILL
        celda_etiqueta.alignment = Alignment(horizontal="right")
        celda_etiqueta.border = _BORDER

        celda_total = ws.cell(row=fila_actual, column=idx_col_total, value=float(df[columna_total].sum()))
        celda_total.font = _TOTAL_FONT
        celda_total.fill = _TOTAL_FILL
        celda_total.number_format = '"$"#,##0'
        celda_total.alignment = Alignment(horizontal="right")
        celda_total.border = _BORDER

    for col_idx, nombre_col in enumerate(df.columns, start=1):
        contenido_max = df[nombre_col].astype(str).map(len).max() if not df.empty else 0
        ancho = max(len(str(nombre_col)), contenido_max) + 4
        ws.column_dimensions[get_column_letter(col_idx)].width = min(ancho, 40)

    ws.freeze_panes = f"A{fila_header + 1}"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
